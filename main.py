import os
import pandas as pd
import pyodbc
from dotenv import load_dotenv
from datetime import datetime
import warnings
import re

# ==============================================================================
#                          CONFIGURAÇÕES GLOBAIS
# ==============================================================================
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
load_dotenv()

# Definição de Caminhos
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PASTA_ENTRADA = os.path.join(BASE_DIR, "RDCs_Originais")
PASTA_SAIDA = os.path.join(BASE_DIR, "Analises")
PASTA_PRONTOS = os.path.join(BASE_DIR, "Prontos_Intranet")

def verificar_pastas():
    """Cria as pastas de entrada e saída caso não existam."""
    for pasta in [PASTA_ENTRADA, PASTA_SAIDA, PASTA_PRONTOS]:
        if not os.path.exists(pasta):
            os.makedirs(pasta)

# ==============================================================================
#                         CONEXÃO COM BANCO DE DADOS
# ==============================================================================
def conectar():
    """Estabelece conexão com SQL Server usando o Driver 17."""
    try:
        driver = "ODBC Driver 17 for SQL Server"
        conn_str = (
            f"DRIVER={{{driver}}};"
            f"SERVER={os.getenv('DB_SERVER')};"
            f"DATABASE={os.getenv('DB_NAME')};"
            f"UID={os.getenv('DB_USER')};"
            f"PWD={os.getenv('DB_PASSWORD')};"
            "TrustServerCertificate=yes;"
        )
        return pyodbc.connect(conn_str)
    except Exception as e:
        print(f"❌ Erro de conexão: {e}")
        return None

# ==============================================================================
#                     INTEGRAÇÃO COM A INTERFACE (APP.PY)
# ==============================================================================
def rodar_automacao_v2(venda_ini, venda_fim, ent_ini, ent_fim, lojas, prazo_t15, arquivos):
    global d_venda_ini, d_venda_fim, T18_VALOR, d_ent_ini, d_ent_fim, T15_VALOR, LOJAS_ALVO, ARQUIVOS_SELECIONADOS

    d_venda_ini = venda_ini
    d_venda_fim = venda_fim
    T18_VALOR = abs((d_venda_fim - d_venda_ini).days) + 1
    d_ent_ini = ent_ini
    d_ent_fim = ent_fim
    T15_VALOR = prazo_t15
    LOJAS_ALVO = lojas
    ARQUIVOS_SELECIONADOS = arquivos

    processar()

# ==============================================================================
#                         EXTRAÇÃO DE DADOS DO EXCEL
# ==============================================================================
def extrair_dados_rdc(caminho_arquivo):
    """Lê as abas do arquivo RDC e extrai Referência, Custo, Múltiplo e Fat. Min."""
    dados_abas = []

    try:
        xls = pd.ExcelFile(caminho_arquivo)
        for aba in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=aba, header=None)

            info = {'ref': "", 'custo': 0, 'multiplo': 1, 'fat_min': 0.0}

            # Localização de Referência e Múltiplo (linhas 0-6)
            for index, row in df.head(7).iterrows():
                linha_texto = " ".join([str(x) for x in row.values if pd.notna(x)])
                match_ref = re.search(r'^(\d{4,6})\s*\|', linha_texto)
                if match_ref:
                    info['ref'] = match_ref.group(1)
                if "Multiplo:" in linha_texto:
                    for i, cel in enumerate(row):
                        if "Multiplo:" in str(cel) and i + 1 < len(row):
                            info['multiplo'] = row[i + 1] if pd.notna(row[i + 1]) else 1

            # Fat. Min. está na coluna 2 das linhas de loja (a partir da linha 8)
            for index, row in df.iloc[8:].iterrows():
                val = row.iloc[2] if len(row) > 2 else None
                if pd.notna(val):
                    try:
                        info['fat_min'] = float(str(val).replace('R$', '').replace('.', '').replace(',', '.').strip())
                        break
                    except:
                        continue

            # Custo está na coluna 3 das linhas de loja (a partir da linha 8)
            for index, row in df.iloc[8:].iterrows():
                val = row.iloc[3] if len(row) > 3 else None
                if pd.notna(val):
                    try:
                        info['custo'] = float(str(val).replace(',', '.'))
                        break
                    except:
                        continue

            if info['ref']:
                info['codigos_busca'] = list(set([info['ref']]))
                dados_abas.append(info)

    except Exception as e:
        print(f"⚠️ Erro ao ler abas: {e}")
    return dados_abas

# ==============================================================================
#                          CONSULTA SQL (BANCO)
# ==============================================================================
def executar_sql(conn, info):
    """Consulta Vendas, Saldos e Pendências no Banco de Dados."""
    cursor = conn.cursor()
    codigos_sql = ",".join([f"'{str(c).strip().zfill(5)}'" for c in info['codigos_busca']])
    v_ini, v_fim = d_venda_ini.strftime('%Y%m%d 00:00:00'), d_venda_fim.strftime('%Y%m%d 23:59:59')

    query = f"""
    SELECT 
        LJ.FIL_RAZAO_SOCIAL AS [Loja],
        ISNULL(SUM(BASE.V), 0) AS [Venda],
        ISNULL(SUM(BASE.E), 0) AS [Est.],
        ISNULL(SUM(BASE.P), 0) AS [Pend.]
    FROM (
        SELECT FIL_CODIGO, FIL_RAZAO_SOCIAL FROM FILIAL (NOLOCK) WHERE FIL_CODIGO IN ({LOJAS_ALVO})
    ) AS LJ
    LEFT JOIN (
        SELECT FID, PID, V, E, P, M.MAT_REFERENCIA
        FROM (
            SELECT VEN_FILIAL as FID, VEN_PRODUTO as PID,
                   SUM(CASE WHEN VEN_NATUREZA = 1 THEN VEN_QUANTIDADE ELSE VEN_QUANTIDADE * -1 END) AS V, 0 AS E, 0 AS P
            FROM VENDAS (NOLOCK) WHERE VEN_INATIVA = 0 AND VEN_DATA >= '{v_ini}' AND VEN_DATA <= '{v_fim}'
            GROUP BY VEN_FILIAL, VEN_PRODUTO
            UNION ALL
            SELECT SAL_FILIAL, SAL_PRODUTO, 0, SUM(SAL_SALDO), 0
            FROM SALDOS (NOLOCK) GROUP BY SAL_FILIAL, SAL_PRODUTO
            UNION ALL
            SELECT ITE_FILIAL, ITE_PRODUTO, 0, 0, SUM(ITE_PEDIDA - ITE_ENTREGUE)
            FROM ITENSPEDIDO (NOLOCK) WHERE ITE_STATUS_NOVO = 5 GROUP BY ITE_FILIAL, ITE_PRODUTO
        ) AS U
        INNER JOIN MATERIAIS M (NOLOCK) ON M.MAT_CODIGO = U.PID
        WHERE M.MAT_REFERENCIA IN ({codigos_sql})
    ) AS BASE ON BASE.FID = LJ.FIL_CODIGO
    GROUP BY LJ.FIL_RAZAO_SOCIAL
    ORDER BY LJ.FIL_RAZAO_SOCIAL
    """
    cursor.execute(query)
    res = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]
    cursor.close()
    return res

# ==============================================================================
#                         MOTOR DE PROCESSAMENTO
# ==============================================================================
def processar():
    verificar_pastas()
    print(f"📁 Arquivos selecionados: {ARQUIVOS_SELECIONADOS}")

    conn = conectar()
    if not conn:
        print("❌ Conexão falhou, encerrando.")
        return

    print(f"✅ Conexão OK")

    for arquivo in ARQUIVOS_SELECIONADOS:
        print(f"🚀 Processando: {arquivo}")
        caminho = os.path.join(PASTA_ENTRADA, arquivo)

        if not os.path.exists(caminho):
            print(f"❌ Arquivo não encontrado: {caminho}")
            continue

        dados_rdc = extrair_dados_rdc(caminho)
        print(f"📊 Abas extraídas: {len(dados_rdc)}")
        lista_final = []

        for item in dados_rdc:
            print(f"🔍 Ref.: {item['ref']} | Fat. Min.: {item['fat_min']} | Custo: {item['custo']} | Múltiplo: {item['multiplo']}")
            resultados = executar_sql(conn, item)
            print(f"📦 Resultados SQL: {len(resultados)} linhas")
            for r in resultados:
                r.update({
                    'Ref.': item['ref'],
                    'Custo': item['custo'],
                    'Qtd. / caixa': item['multiplo'],
                    'Fat_Min_RDC': item['fat_min']
                })
                lista_final.append(r)

        if lista_final:
            df = pd.DataFrame(lista_final)
            if 'Fat_Min_RDC' not in df.columns:
                df['Fat_Min_RDC'] = 0

            df = df.sort_values(by=["Ref.", "Loja"])

            # Preparação das colunas de fórmulas
            cols_formulas = ["Cob.", "Ped.", "Cob. máx.", "Cob. ent.", "Est. ent.", "Q1", "Q2", "R1", "R2", "T1", "T2"]
            for col in cols_formulas:
                df[col] = ""

            ordem = ["Ref.", "Custo", "Qtd. / caixa", "Loja", "Venda", "Est.", "Pend.",
                     "Cob.", "Ped.", "Cob. máx.", "Cob. ent.", "Est. ent.", "Q1", "Q2", "R1", "R2", "T1", "T2"]
            fat_exibicao = df['Fat_Min_RDC'].iloc[0] if 'Fat_Min_RDC' in df.columns else 0  # ← LIDO ANTES de df = df[ordem]
            df = df[ordem]

            # --- GERAÇÃO DO EXCEL FORMATADO ---
            nome_sem_ext = os.path.splitext(arquivo)[0]
            base = os.path.join(PASTA_SAIDA, f"Analise_{nome_sem_ext}.xlsx")
            caminho_saida = base
            n = 1
            while os.path.exists(caminho_saida):
                caminho_saida = os.path.join(PASTA_SAIDA, f"Analise_{nome_sem_ext} ({n}).xlsx")
                n += 1
            try:
                writer = pd.ExcelWriter(caminho_saida, engine='xlsxwriter')
                df.to_excel(writer, index=False, sheet_name='Analise', startrow=0)
                workbook  = writer.book
                worksheet = writer.sheets['Analise']
                worksheet.freeze_panes(1, 7)

                # Definição de Formatos Excel
                fmt_amarelo   = workbook.add_format({'bg_color': '#FFFF00', 'border': 1, 'align': 'center', 'bold': True})
                fmt_azul      = workbook.add_format({'bg_color': '#0070C0', 'font_color': 'white', 'border': 1, 'align': 'center', 'bold': True})
                fmt_bold      = workbook.add_format({'bold': True, 'border': 1, 'align': 'right'})
                fmt_money_y   = workbook.add_format({'num_format': 'R$ #,##0.00', 'bg_color': '#FFFF00', 'border': 1})
                fmt_money     = workbook.add_format({'num_format': 'R$ #,##0.00', 'border': 1})
                fmt_date      = workbook.add_format({'num_format': 'dd/mm/yyyy', 'bg_color': '#FFFF00', 'border': 1, 'align': 'center'})
                fmt_int       = workbook.add_format({'num_format': '0', 'border': 1, 'align': 'center'})
                fmt_linha_sep = workbook.add_format({'bottom': 2, 'bottom_color': '#000000'})

                # Painel de Controle Lateral no Excel
                col_painel   = 19
                hoje_geracao = datetime.now().date()

                worksheet.write(10, col_painel,   "Gerado em:",   fmt_bold)
                worksheet.write(10, col_painel+1, hoje_geracao,   fmt_date)
                worksheet.write(11, col_painel,   "Fat. Mín.:",   fmt_bold)
                worksheet.write(11, col_painel+1, fat_exibicao,   fmt_money_y)
                worksheet.write(12, col_painel,   "Entrega de:",  fmt_bold)
                worksheet.write(12, col_painel+1, d_ent_ini.date(), fmt_date)
                worksheet.write(13, col_painel,   "Entrega até:", fmt_bold)
                worksheet.write(13, col_painel+1, d_ent_fim.date(), fmt_date)
                worksheet.write(14, col_painel,   "Prazo máx.:",  fmt_bold)
                worksheet.write(14, col_painel+1, T15_VALOR,      fmt_amarelo)
                worksheet.write(15, col_painel,   "Venda de:",    fmt_bold)
                worksheet.write(15, col_painel+1, d_venda_ini.date(), fmt_date)
                worksheet.write(16, col_painel,   "Venda até:",   fmt_bold)
                worksheet.write(16, col_painel+1, d_venda_fim.date(), fmt_date)
                worksheet.write(17, col_painel,   "Período:",     fmt_bold)
                worksheet.write(17, col_painel+1, T18_VALOR,      fmt_amarelo)

                # Escrita de Fórmulas Dinâmicas
                for i in range(len(df)):
                    row = i + 1
                    idx = i + 2
                    worksheet.write(row, 1, df.iloc[i]['Custo'], fmt_money)
                    worksheet.write(row, 2, df.iloc[i]['Qtd. / caixa'])
                    worksheet.write_formula(row, 7,  f'=IFERROR((F{idx}+G{idx})/E{idx}*{T18_VALOR},"-")')
                    worksheet.write_formula(row, 8,  f'=SUM(M{idx}:N{idx})', fmt_azul)
                    worksheet.write_formula(row, 9,  f'=IFERROR((F{idx}+G{idx}+I{idx})/E{idx}*{T18_VALOR},"-")', fmt_int)
                    worksheet.write_formula(row, 10, f'=IFERROR(((F{idx}+G{idx}+I{idx})/E{idx}*{T18_VALOR})-{T15_VALOR},"-")', fmt_int)
                    worksheet.write_formula(row, 11, f'=IFERROR((F{idx}+G{idx}+I{idx})-(E{idx}*{T15_VALOR}/{T18_VALOR}),"-")', fmt_int)
                    worksheet.write(row, 12, 0, fmt_amarelo)
                    worksheet.write(row, 13, 0, fmt_amarelo)
                    worksheet.write_formula(row, 14, f'=M{idx}*B{idx}', fmt_money)
                    worksheet.write_formula(row, 15, f'=N{idx}*B{idx}', fmt_money)
                    worksheet.write_formula(row, 16, f'=SUMIFS(O:O,D:D,D{idx})', fmt_money)
                    worksheet.write_formula(row, 17, f'=SUMIFS(P:P,D:D,D{idx})', fmt_money)

                    # Linha separadora visual por Referência
                    if i < len(df) - 1 and df.iloc[i]['Ref.'] != df.iloc[i+1]['Ref.']:
                        worksheet.conditional_format(row, 0, row, 17, {'type': 'no_errors', 'format': fmt_linha_sep})

                writer.close()
                print(f"✅ Análise finalizada e salva: {arquivo}")
            except Exception as e:
                print(f"❌ Erro ao salvar Excel: {e}")
        else:
            print(f"⚠️ Nenhum dado encontrado para: {arquivo}")

    conn.close()

# ==============================================================================
#                     PREENCHIMENTO DO RDC COM Q1
# ==============================================================================
def preencher_rdc_com_q1(caminho_analise, caminho_rdc):
    """Lê Q1 da análise e preenche a coluna K do RDC original."""
    try:
        # Lê a análise
        df_analise = pd.read_excel(caminho_analise, sheet_name='Analise', header=0)
        df_analise['Ref.'] = df_analise['Ref.'].astype(str).str.strip()
        df_analise['Loja'] = df_analise['Loja'].astype(str).str.strip()

        # Abre o RDC com openpyxl para editar mantendo formatação
        from openpyxl import load_workbook
        wb = load_workbook(caminho_rdc)

        for aba in wb.sheetnames:
            ws = wb[aba]

            # Filtra as linhas da análise para esta referência
            df_ref = df_analise[df_analise['Ref.'] == aba.strip()]
            if df_ref.empty:
                print(f"⚠️ Aba {aba} sem dados na análise, pulando.")
                continue

            # Percorre as linhas de loja no RDC (a partir da linha 9 = índice 8)
            for row in ws.iter_rows(min_row=9):
                loja_rdc = str(row[1].value).strip() if row[1].value else ""
                if not loja_rdc:
                    continue

                # Busca o Q1 correspondente na análise
                match = df_ref[df_ref['Loja'].str.strip() == loja_rdc]
                if not match.empty:
                    q1_valor = match.iloc[0]['Q1']
                    row[10].value = q1_valor  # coluna K = índice 10
                    print(f"✅ {aba} | {loja_rdc} → Q1={q1_valor}")
                else:
                    print(f"⚠️ {aba} | {loja_rdc} → não encontrado na análise")

        # Salva o RDC preenchido na pasta Prontos_Intranet
        nome_sem_ext = os.path.splitext(os.path.basename(caminho_rdc))[0]
        base = os.path.join(PASTA_PRONTOS, f"RDC_Preenchido_{nome_sem_ext}.xlsx")
        caminho_saida = base
        n = 1
        while os.path.exists(caminho_saida):
            caminho_saida = os.path.join(PASTA_PRONTOS, f"RDC_Preenchido_{nome_sem_ext} ({n}).xlsx")
            n += 1
        wb.save(caminho_saida)
        print(f"✅ RDC preenchido salvo: {caminho_saida}")
        return True

    except Exception as e:
        print(f"❌ Erro ao preencher RDC: {e}")
        return False

# ==============================================================================
#                            INICIALIZAÇÃO
# ==============================================================================
if __name__ == "__main__":
    print("Módulo de lógica RDC carregado com sucesso.")
    print("Para utilizar a interface visual, execute o arquivo 'app.py'.")